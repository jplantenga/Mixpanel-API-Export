import os
import json
import base64
import requests
import pandas as pd
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load environment variables from .env file
def load_env_vars():
    print("Loading environment variables...")
    load_dotenv()
    print("Environment variables loaded successfully.")

# Prepare API headers
def prepare_headers(api_secret):
    if not api_secret:
        raise ValueError("API Secret cannot be empty.")
    print("Preparing headers for API request...")
    encoded_secret = base64.b64encode(f"{api_secret}:".encode('utf-8')).decode('utf-8')
    return {
        "Authorization": f"Basic {encoded_secret}",
        "accept": "text/plain"
    }

# Fetch data from Mixpanel API
def fetch_data_from_api(base_url, endpoint, headers, params):
    print("Fetching data from Mixpanel API...")
    try:
        response = requests.get(f"{base_url}{endpoint}", headers=headers, params=params, stream=True)
        response.raise_for_status()
        print(f"API Response Code: {response.status_code}")
        return response
    except requests.RequestException as e:
        print(f"An error occurred while fetching data from Mixpanel: {e}")
        return None

# Write API response to a temporary file
def write_response_to_temp_file(response, temp_file_path):
    print("Writing API response to temporary file...")
    with open(temp_file_path, "wb") as f:
        try:
            for chunk in response.iter_content(chunk_size=128):
                f.write(chunk)
        except IOError as e:
            print(f"An I/O error occurred: {e}")

# Read and process events from the temporary file
def read_and_process_events(temp_file_path):
    print("Reading and processing events from temporary file...")
    events = []
    with open(temp_file_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                print("Skipping empty lines")
                continue
            try:
                event = json.loads(line)
                events.append(event)
            except json.JSONDecodeError as e:
                print(f"Could not decode a line as JSON: {e}")
    return events

# Validate date format
def validate_date(date_str):
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        raise ValueError("Incorrect date format, should be YYYY-MM-DD")

# Adjust column widths in Excel sheet
def adjust_column_widths(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

if __name__ == "__main__":
    # Load environment variables
    load_env_vars()
    API_SECRET = os.getenv("MIXPANEL_API_SECRET")
    if not API_SECRET:
        raise EnvironmentError("Missing environment variable: MIXPANEL_API_SECRET")

    # Basic settings for API call
    BASE_URL = "https://data-eu.mixpanel.com/api/2.0/"
    EVENTS_ENDPOINT = "export"

    HEADERS = prepare_headers(API_SECRET)

    # Collect and validate date input
    from_date = input("Enter a start date (YYYY-MM-DD): ")
    to_date = input("Enter an end date (YYYY-MM-DD): ")
    validate_date(from_date)
    validate_date(to_date)

    # Make API call and receive response
    PARAMS = {'from_date': from_date, 'to_date': to_date}
    response = fetch_data_from_api(BASE_URL, EVENTS_ENDPOINT, HEADERS, PARAMS)

    if response:
        import tempfile
        TEMP_FILE_PATH = tempfile.NamedTemporaryFile(delete=False).name
        write_response_to_temp_file(response, TEMP_FILE_PATH)
        events = read_and_process_events(TEMP_FILE_PATH)
        
        # Process received events and save as CSV and Excel
        if events:  # If there are events
            df = pd.json_normalize(events)

            # Write to CSV
            csv_file_path = "/path/to/your/csv/file.csv"
            df.to_csv(csv_file_path, index=False)
            print(f"Events saved in CSV file: {csv_file_path}")

            # Write to Excel
            excel_file_path = "/path/to/your/excel/file.xlsx"
            df.to_excel(excel_file_path, index=False)
            print(f"Events saved in Excel file: {excel_file_path}")

            # Adjust the column width
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            adjust_column_widths(worksheet)
            workbook.save(excel_file_path)
            print("Column widths have been adjusted in the Excel file.")
        else:  # If there are no events
            print("No events found.")
